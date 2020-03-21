# Open Excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# To iterate files
import os

# To parse CSV
import csv

# Convert latin-1 to unicode
def unicode_csv_reader(utf8_data, dialect=csv.excel, **kwargs):
    csv_reader = csv.reader(utf8_data, dialect=dialect, **kwargs)
    for row in csv_reader:
        yield [unicode(cell, 'latin-1') for cell in row]

# Root directory to iterate over files
dirName = 'csvFiles'

# Recurse over root folder
for path, subdirs, files in os.walk(dirName):
    # Iterate through sub-directories
    for subdir in subdirs:
        print subdir
        wb = Workbook()
        basePath = os.path.join(path,subdir)
        # Get all files in directory
        files = [f for f in sorted(os.listdir(os.path.join(basePath))) \
                 if os.path.isfile(os.path.join(basePath,f))]
        files.sort(key=lambda f: int(filter(str.isdigit, f)))
        # Iterate through files
        for f in files:
            # Create sheets and export csv contents to it
            print f
            ws = wb.create_sheet(f)
            # Read CSV file contents
            # Call this if CSV is latin-1 encoding
            #reader = unicode_csv_reader(open(os.path.join(basePath, f)))
            reader = csv.reader(open(os.path.join(basePath, f)))
            for row in reader:
                ws.append(row)
            ws['B1'] = 'Enter Word'
            #format_cell_range(worksheet, '1', fmtHeader)
            black_format = openpyxl.styles.colors.Color(rgb='000000')
            hide_fill = openpyxl.styles.fills.PatternFill(patternType='solid', bgColor=black_format)
            for rows in ws.iter_cols(min_col=1, max_col=1, min_row=2):
                for cell in rows:
                    cell.fill = hide_fill
            green_format = openpyxl.styles.colors.Color(rgb='00FF00')
            green_fill = openpyxl.styles.fills.PatternFill(patternType='solid', bgColor=green_format)
            ws.conditional_formatting.add('B2:B151', FormulaRule(formula=['A2=B2'], fill=green_fill))
        # Delete default(first) sheet
        wb.remove(wb['Sheet'])
        # Save
        print 'saving file = ' + subdir+'.xlsx'
        wb.save( subdir + '.xlsx')


